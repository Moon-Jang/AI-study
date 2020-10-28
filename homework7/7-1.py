import numpy as np
import random


def act_tlu(x):
    if x > 0.0:
        return 1.0
    return 0.0


def act_sigmoid(x):
    return 1.0 / (1.0 + np.exp(-x))


def get_error(target, b):
    return target - b


x = np.array([[1.0, 0.0, 0.0], [1.0, 0.0, 1.0], [1.0, 1.0, 0.0], [1.0, 1.0, 1.0]])
t = np.array([0.0, 0.0, 0.0, 1.0])

lrate = 0.1

w = np.zeros(3)

w[0] = random.random()
w[1] = random.random()
w[2] = random.random()

for epoch in range(100):
    # print("epoch ", epoch)
    for i in range(4):
        out = w[0] * x[i][0] + w[1] * x[i][1] + w[2] * x[i][2]
        b = act_tlu(out)
        error = get_error(t[i], b)
        for j in range(3):
            w[j] = w[j] + lrate * error * x[i][j]

from openpyxl import Workbook

write_wb = Workbook()
write_ws = write_wb.active

input = [1.0, 0.0, 0.0]
for i in range(11):
    input[1] = round(i * 0.1, 1)
    write_ws.cell(1, i + 2, round(i * 0.1, 1))
    write_ws.cell(i + 2, 1, round(i * 0.1, 1))
    for j in range(11):
        input[2] = round(j * 0.1, 1)
        out = w[0] * input[0] + w[1] * input[1] + w[2] * input[2]
        result = act_sigmoid(out)
        print(input[1], input[2], result)
        write_ws.cell(i + 2, j + 2, result)
    input[2] = 0.0

write_wb.save('7_1_graph.xlsx')
